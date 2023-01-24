from flask import Flask, redirect, render_template, request
import pandas as pd
import openpyxl
from openpyxl import load_workbook

app = Flask(__name__)

Slots = [1,2,3,4,5]
Sheets = ['ADA','CN','PDS','SE','PE','PLSD','IPDC','DE-II-A']


@app.route("/", methods=["GET", "POST"])
def index():
    error = []
    #If request is POST then taking form Data and working on them
    if request.method == 'POST':

        #Storing and Validating data
        date = request.form.get('date')
        print(date)
        if date == '':
            error.append('Please Enter Date!!')
            return render_template('fail.html',message = error),400


        sheet = request.form.get('sheet')
        if sheet == None:
            error.append('Please Enter Sheet Name!!')
            return render_template('fail.html',message = error),400
        elif sheet not in Sheets:
            error.append('Please Enter Valid Sheet Name!!')
            return render_template('fail.html',message = error),400

        slot = request.form.get('slot')
        print(slot)
        if slot == None:
            error.append('Please Enter Slot Number')
            return render_template('fail.html',message = error),400
        if int(slot) not in Slots:
            error.append('Please Enter Valid Slot Number')
            return render_template('fail.html',message = error),400

        #Opening the current attendence sheet
        attendence_file = request.files['file']
        filename = str(attendence_file.filename)

        if filename == '':
            error.append('Please Upload Attendence File!!')
            return render_template('fail.html',message = error),400
        fileformat = filename[len(filename)-5] + filename[len(filename)-4] + filename[len(filename)-3] + filename[len(filename)-2] + filename[len(filename)-1]
        if fileformat != '.xlsx':
            error.append('Please Upload .xlsx Extention File!!')
            return render_template('fail.html',message = error),400

        data_xls = pd.read_excel(attendence_file)


        #Opening the Cumulitive Musturd
        mustard = load_workbook("Mustered.xlsx")
        active_sheet = mustard[sheet]


        #Creating Dictionary of Presenties
        attendes = {}
        itr=0
        while True:
            itr += 1
            try:
                if data_xls['Name'][itr] != None and int(data_xls['Duration'][itr]) > 40:
                    attendes[data_xls['Name'][itr]] = int(data_xls['Duration'][itr])
                elif data_xls['Name'][itr] != None and int(data_xls['Duration'][itr]) <= 40:
                    continue
            except:
                break

        print(attendes)


        #Iterating through mustard to find last empty row
        r_counter = 0
        while True:
            r_counter += 1
            if active_sheet.cell(4,r_counter).value == None:
                break
        print(r_counter)


        #updating date and slot in main Mustard
        active_sheet.cell(4,r_counter).value = date
        active_sheet.cell(5,r_counter).value = slot
        print(active_sheet.cell(4,r_counter).value)
        print(active_sheet.cell(5,r_counter).value)


        #Iterating through mustard to find last empty column and updating it with
            # attendence of presenties with the help of attendes Dictionary
        c_counter = 6
        while True:
            c_counter += 1
            if active_sheet.cell(c_counter,2).value == None:
                break
            elif active_sheet.cell(c_counter,2).value in attendes:
                active_sheet.cell(c_counter,r_counter).value = 'P'
            else:
                active_sheet.cell(c_counter,r_counter).value = 'A'
            print(active_sheet.cell(c_counter,r_counter).value)


        mustard.save(filename="Mustered.xlsx")

        return render_template('success.html')
    else:
        return render_template('index.html',slots = Slots,sheets=Sheets)